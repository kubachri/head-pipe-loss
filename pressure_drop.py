#import libraries
import pandas as pd
#import clear_output
from IPython.display import clear_output
#import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
#import counter for merging dictionaries
from collections import Counter

wb=Workbook()
filepath=r'C:\Users\chris\Documents\Python Projects\head loss\output\Pressure_Drop_Data.xlsx'

#open piping data file
pipe_reducers=pd.read_excel(r"C:\Users\chris\Documents\Python Projects\head loss\data\Piping Data.xlsx",sheet_name='Reducers')

#convert data file to DataFrame
df=pd.DataFrame(pipe_reducers)

#create the reducers' z-value dictionary:
reducer_dict=dict(df.values)

#PPR Pipe Dimension, Fitting Type and Zeta Values 
ppr_dimension={20:14.4,25:18,32:26.2,40:32.6,50:40.8,63:51.4,75:61.4,90:73.6,110:90,125:102.2}
ppr_fittings=['Socket','Elbow90','Elbow45','Tee','Reducer']
z_socket=0.25
z_elbow90=1.2
z_elbow45=0.5
z_tee=1.2

#Create Excel File for the Output
excel=wb.active
#Initializing Headers
excel["A1"].value='Node'
excel["A3"].value='Pipe'
r=range(5,6+2*len(ppr_fittings),2)
j=0
for i in r:
	if i==r[-1]:
		excel.cell(row=i,column=1).value='Pressure Loss (mH2O)'
	else:
		excel.cell(row=i,column=1).value=ppr_fittings[j]
		j+=1


#CLASSES:

#Class of Pipe characteristics
class PipePPR:
	type='Pipe'

	def __init__(self):
		while True:
			try:
				self.diameter=int(input('Enter Pipe Diameter (in mm): '))
				break
			except:
				print('Invalid Input. Try again!')
		while self.diameter not in ppr_dimension.keys():
			while True:
				try:
					self.diameter=int(input('Invalid Diameter. Enter Pipe Diameter (in mm): '))
					break
				except:
					print('Invalid diameter.')
		self.hydraulic_diam=ppr_dimension[self.diameter]
		while True:
			try:
				self.length=float(input('Enter pipe length (in m): '))
				break
			except:
				print('Length must be a number!')
		self.hydraulic_diam=ppr_dimension[self.diameter]
		self.name='PP-R Φ'+str(self.diameter)

	def __str__(self):
		return self.name

	def fi(self):
		return self.name.split()[1]


#Class of Fitting Characteristics
class FittingPPR:

	def __init__(self):
		self.type=input('Enter the type of Fitting (Tee,Elbow90,Elbow45,Reducer,Socket): ')
		self.type=self.type.replace(" ","")
		self.type=self.type.lower()
		self.type=self.type.capitalize()

		#Check for Correct fitting type input
		while self.type.capitalize() not in ppr_fittings:
			self.type=input('You have to type on of the following (Tee,Elbow90,Elbow45,Reducer,Socket): ')
			self.type=self.type.replace(" ","")
			self.type=self.type.lower()
			self.type=self.type.capitalize()

		#check for correct diameter input
		while True:
			try:
				self.diameter=int(input('Enter Fitting Diameter (in mm): '))
				break
			except:
				print('Invalid Input. Try again!')
		while self.diameter not in ppr_dimension.keys():
			self.diameter=int(input('Invalid Diameter. Enter Fitting Diameter (in mm): '))
		self.hydraulic_diam=ppr_dimension[self.diameter]

		#print styling
		self.name=self.type.capitalize()+' Φ'+str(self.diameter)

		#Fitting's zeta-value
		if self.type.capitalize()=='Reducer':
			while True:
				try:
					reduced=input('Enter the reduced diameter (in mm):')
					break
				except:
					print('Invalid diameter input!')
			self.name=self.name+'x'+reduced
			self.zeta_value=reducer_dict[str(self.diameter)+'x'+reduced]
		elif self.type.capitalize()=='Tee':
			self.zeta_value=1.2
		elif self.type.capitalize()=='Socket':
			self.zeta_value=0.25
		elif self.type.capitalize()=='Elbow90':
			self.zeta_value=1.2
		elif self.type.capitalize()=='Elbow45':
			self.zeta_value=0.5

	#Print Statements
	def __str__(self):
		return self.name

	def fi(self):
		return self.name.split()[1]

#PrintFormat Text Underline Format (Format.underline + 'text' + Format.end)
class Format:
    end = '\033[0m'
    underline = '\033[4m'

#FUNCTIONS:

#Hanzen-Williams 
def pipe_loss(pipe,flow_rate):

	#losses per 100m pipe
	f_friction=((100/150)**1.852)*(((flow_rate/3600*1000)*15.852)**1.852)/((pipe.hydraulic_diam*0.03937)**4.8655)*0.2083*304.8*3.28

	#total pipe losses (in mH2O)
	head_loss_pipe=f_friction*pipe.length/100000

	#output of friction losses in pipe
	return head_loss_pipe


def fitting_loss(flow_rate,fitting):
	velocity = ((flow_rate/3600*1000)/1000)/(3.14*((fitting.hydraulic_diam/1000/2)**2))
	head_loss_fitting=fitting.zeta_value*velocity**2*1000*0.5/1000/9.81

	return head_loss_fitting

#function for rounding up to 2 decimals
def roundupnum(num):
	return round(num,2)

#fucntion for doubling the parts (supply-return)
def doubled(value):
	return value*2


# Excel Styling
def excel_style(worksheet, max_row, max_column):
	# Excel zero values to none
	for row in range(1, last_row + 1):
		for column in range(last_column - 2, last_column + 1):
			if excel.cell(row=row, column=column).value == 0:
				excel.cell(row=row, column=column).value = None

	# Bold Headers
	for column in excel["A:A"]:
		try:
			if column.value[0] == 'Φ':
				continue
		except TypeError:
			pass
		column.font = Font(bold=True)
	for row in excel["1:1"]:
		row.font = Font(bold=True)


# ~~~~~~~~~ MODEL START ~~~~~~~~~

#Instructions
print("""
	In every node-i, the length of pipe is calculated from node-i to node-i+1. 
	The fittings are calculated with the same logic.
	""")
#flag
calc=True

#initializing variables
dp_branch=0

#Loop for node calculations
while calc==True:
	while True:
		try:
			node_num=int(input('Please insert the number of nodes: '))
			break
		except:
			print('The number of nodees must be integer!')		

	#Total Pressure Drop for Each Node
	node_dp = {}

	#Total Parts for Each Node
	node_parts={}	

	#Total Parts
	total_parts={}

	#Iteration for each node
	for node in range(node_num):
		print (f'\n node {node}\n')	
		node_dp_sum=0

		#Node Flow Rate
		while True:
			try:
				q=float(input('Flow Rate (m3/h) for this node: '))
				break
			except:
				print('Invalid Input. Format must be: number.number')
		
		#Node Pipe Length
		pipe=PipePPR()
		dp_pipe=pipe_loss(pipe,q)

		# Support Index Flag for Cell Position in Excel
		idx = 0

		# Excel Write
		for i in range(1, len(excel["A"]) + 1):
			if excel["A" + str(i)].value == 'Pipe':
				idx = i + 1
				for row in range(idx, len(excel["A"])):
					if excel["A" + str(row)].value == pipe.fi():
						excel.cell(row=row, column=2 + node).value = pipe.length
						break
					try:
						if excel["A" + str(row)].value[0] != 'Φ':
							excel.insert_rows(idx=i + 1, amount=1)
							excel.cell(row=i + 1, column=1).value = pipe.fi()
							excel.cell(row=i + 1, column=2 + node).value = pipe.length
							break
					except:
						pass

		node_parts[node]={pipe.type:pipe.fi()}
		node_parts[node][pipe.type]={pipe.fi():pipe.length}
		node_dp_sum=node_dp_sum+dp_pipe

		#Node Fittings
		while True:
			try:
				fittings_num=int(input('\nEnter the number of fittings for this node: '))
				break
			except:
				print('\nNumber must be an integer!')
		
		for part in range(fittings_num):
			fitting=FittingPPR()
			if fitting.type in node_parts[node].keys():
				if fitting.fi() in node_parts[node][fitting.type].keys():
					node_parts[node][fitting.type][fitting.fi()]+=1
				else:
					node_parts[node][fitting.type].update({fitting.fi():1})
			else:
				node_parts[node].update({fitting.type:{fitting.fi():1}})

			#Total Parts
			if fitting.type in total_parts.keys():
				if fitting.fi() in total_parts[fitting.type].keys():
					total_parts[fitting.type][fitting.fi()]+=1
				else:
					total_parts[fitting.type].update({fitting.fi():1})
			else:
				total_parts.update({fitting.type:{fitting.fi():1}})


			dp_fitting=fitting_loss(q,fitting)
			node_dp_sum=node_dp_sum+dp_fitting

		#Pressure Drop		
		node_dp[node]=node_dp_sum
		dp_branch=dp_branch+node_dp_sum

	calc=False

	#rounding dp to 2 decimals
	for key in node_dp:
		node_dp[key]=roundupnum(node_dp[key])

	#clear program text before output
	clear_output()

	#OUTPUT
	print('\n')
	print(Format.underline+'NODE PARTS'+Format.end)

	#Node Parts Output
	for node in node_parts:
		print('\n')
		print ('\t'+Format.underline+f'Node {node}'+Format.end+': ')
		for fittings in node_parts[node].keys():
			print('\n')
			print('\t\t'+Format.underline+fittings+Format.end)
			for x,y in node_parts[node][fittings].items():
				print(f'\t\t{x}: {y}')
				
	print('\n')
	print(Format.underline+'PRESSURE DROP'+Format.end)
	print('\n')

	#Node Pressure Drop Output
	for node in node_dp:
		print(f'\tNode {node}: dP={node_dp[node]} m(H) \n')
	
	#Print Total Branch Pressure
	print(Format.underline+'The Total Pressure Drop (Supply+Return) for this Branch is'+Format.end+f' : {doubled(roundupnum(dp_branch))} m(H)')

	#Print Total Branch Parts
	for key in total_parts.keys():
		for value in total_parts[key]:
			total_parts[key]=doubled(total_parts[key][value])

	print(Format.underline+'The list of Fittings (Supply+Return) for this Branch are'+Format.end+f' : {total_parts}')
